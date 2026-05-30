"""Generate a polished report from marcelo_final.xlsx."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

SRC = r"C:\Users\marce\OneDrive\Documentos\marcelo_final.xlsx"
DST = r"C:\Users\marce\OneDrive\Documentos\marcelo_final_report.xlsx"

# Schema (dataclass field order) — 46 columns total
COLS = [
    "experiment_id", "username", "notes", "approach", "inference_model",
    "inference_model_size", "debug", "mode", "transcription_model", "is_rag",
    "embedding_model", "embedding_model_size",
    "mean_question_accuracy", "mean_time", "mean_search_time",
    "mean_reasoning_time", "mean_transcription_time",
    "ent_time", "ent_acc", "ent_search", "ent_reason", "ent_trans",
    "sci_time", "sci_acc", "sci_search", "sci_reason", "sci_trans",
    "anc_time", "anc_acc", "anc_search", "anc_reason", "anc_trans",
    "math_time", "math_acc", "math_search", "math_reason", "math_trans",
    "news_time", "news_acc", "news_search", "news_reason",
    "phil_time", "phil_acc", "phil_search", "phil_reason",
    "timestamp",
]

df = pd.read_excel(SRC, header=None, names=COLS)
df["label"] = df["notes"]

# Styles
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
BAND_FILL = PatternFill("solid", start_color="F2F2F2")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name=FONT, bold=True, color="1F4E78", size=11)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, start_row, end_row, ncols, band=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c == 1:
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER
            if band and (r - start_row) % 2 == 1:
                cell.fill = BAND_FILL


def auto_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ──────────────────────────── Sheet 1: Resumo ────────────────────────────
ws = wb.active
ws.title = "Resumo Executivo"

ws["A1"] = "Relatório Final — Comparação de Modelos (Millionaire RAG)"
ws["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
ws.merge_cells("A1:H1")
ws["A2"] = f"Embedding: nomic-embed-text  |  Translator: qwen2.5:0.5b  |  Math: qwen2-math:1.5b  |  10 jogos/tema"
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A2:H2")
ws.row_dimensions[1].height = 24

# Sort by overall accuracy desc
summary = df[["label", "inference_model", "mean_question_accuracy",
              "mean_time", "mean_search_time", "mean_reasoning_time"]].copy()
summary["efficiency"] = summary["mean_question_accuracy"] / summary["mean_time"]
summary = summary.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
summary.insert(0, "rank", range(1, len(summary) + 1))

headers = ["#", "Experimento (inference/fallback)", "Inference Model",
           "Acurácia Geral", "Tempo Médio (s)", "Tempo Busca (s)",
           "Tempo Reasoning (s)", "Eficiência (acc/s)"]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))
ws.row_dimensions[HEADER_ROW].height = 30

start = HEADER_ROW + 1
for i, row in summary.iterrows():
    r = start + i
    ws.cell(row=r, column=1, value=int(row["rank"]))
    ws.cell(row=r, column=2, value=row["label"])
    ws.cell(row=r, column=3, value=row["inference_model"])
    ws.cell(row=r, column=4, value=float(row["mean_question_accuracy"]))
    ws.cell(row=r, column=5, value=float(row["mean_time"]))
    ws.cell(row=r, column=6, value=float(row["mean_search_time"]))
    ws.cell(row=r, column=7, value=float(row["mean_reasoning_time"]))
    ws.cell(row=r, column=8, value=float(row["efficiency"]))

end = start + len(summary) - 1
style_body(ws, start, end, len(headers))

for r in range(start, end + 1):
    ws.cell(row=r, column=4).number_format = "0.0%"
    for c in (5, 6, 7):
        ws.cell(row=r, column=c).number_format = "0.00"
    ws.cell(row=r, column=8).number_format = "0.000"

# Color scales
ws.conditional_formatting.add(
    f"D{start}:D{end}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))
ws.conditional_formatting.add(
    f"E{start}:E{end}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="F8696B"))
ws.conditional_formatting.add(
    f"H{start}:H{end}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

auto_width(ws, [5, 32, 18, 16, 16, 16, 18, 18])

# Key findings
fr = end + 3
ws.cell(row=fr, column=1, value="Principais Conclusões").font = Font(name=FONT, bold=True, size=13, color="1F4E78")
ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=8)

best = summary.iloc[0]
fastest = summary.sort_values("mean_time").iloc[0]
most_eff = summary.sort_values("efficiency", ascending=False).iloc[0]

bullets = [
    f"• Melhor acurácia geral: {best['label']} ({best['mean_question_accuracy']:.1%}) em {best['mean_time']:.2f}s por questão.",
    f"• Mais rápido: {fastest['label']} ({fastest['mean_time']:.2f}s/questão, acurácia {fastest['mean_question_accuracy']:.1%}).",
    f"• Maior eficiência (acc/s): {most_eff['label']} — {most_eff['efficiency']:.3f}.",
    f"• Combinações com gemma3:4b como inference dominam o topo da tabela.",
    f"• O fallback influencia mais a acurácia que o tempo: trocar o fallback de gemma3:4b alterou a acurácia em {(df[df['inference_model']=='gemma3:4b']['mean_question_accuracy'].max()-df[df['inference_model']=='gemma3:4b']['mean_question_accuracy'].min()):.1%}.",
]
for i, b in enumerate(bullets):
    ws.cell(row=fr + 1 + i, column=1, value=b).font = BODY_FONT
    ws.merge_cells(start_row=fr + 1 + i, start_column=1, end_row=fr + 1 + i, end_column=8)

ws.freeze_panes = "A5"

# ──────────────────────────── Sheet 2: Acurácia por Tema ─────────────────
ws2 = wb.create_sheet("Acurácia por Tema")
ws2["A1"] = "Acurácia por Tema"
ws2["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
ws2.merge_cells("A1:I1")

theme_cols = [("Entertainment", "ent_acc"), ("Science & Nature", "sci_acc"),
              ("Ancient History", "anc_acc"), ("Maths", "math_acc"),
              ("News", "news_acc"), ("Philosophy & Psych.", "phil_acc")]
headers2 = ["Experimento", "Inference"] + [t[0] for t in theme_cols] + ["Média Geral"]
HR = 3
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=HR, column=i, value=h)
style_header(ws2, HR, len(headers2))
ws2.row_dimensions[HR].height = 30

df_acc = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
for i, row in df_acc.iterrows():
    r = HR + 1 + i
    ws2.cell(row=r, column=1, value=row["label"])
    ws2.cell(row=r, column=2, value=row["inference_model"])
    for j, (_, col) in enumerate(theme_cols, start=3):
        ws2.cell(row=r, column=j, value=float(row[col]))
    ws2.cell(row=r, column=3 + len(theme_cols), value=float(row["mean_question_accuracy"]))

end2 = HR + len(df_acc)
style_body(ws2, HR + 1, end2, len(headers2))
for r in range(HR + 1, end2 + 1):
    for c in range(3, len(headers2) + 1):
        ws2.cell(row=r, column=c).number_format = "0.0%"

# Color each theme column individually
for j in range(3, len(headers2) + 1):
    col_letter = get_column_letter(j)
    ws2.conditional_formatting.add(
        f"{col_letter}{HR+1}:{col_letter}{end2}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))

auto_width(ws2, [32, 18] + [16] * len(theme_cols) + [16])
ws2.freeze_panes = "C4"

# ──────────────────────────── Sheet 3: Tempos por Tema ───────────────────
ws3 = wb.create_sheet("Tempos por Tema")
ws3["A1"] = "Tempo Médio (s) por Tema"
ws3["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
ws3.merge_cells("A1:I1")

time_cols = [("Entertainment", "ent_time"), ("Science & Nature", "sci_time"),
             ("Ancient History", "anc_time"), ("Maths", "math_time"),
             ("News", "news_time"), ("Philosophy & Psych.", "phil_time")]
headers3 = ["Experimento", "Inference"] + [t[0] for t in time_cols] + ["Média Geral"]
for i, h in enumerate(headers3, start=1):
    ws3.cell(row=HR, column=i, value=h)
style_header(ws3, HR, len(headers3))
ws3.row_dimensions[HR].height = 30

df_time = df.sort_values("mean_time").reset_index(drop=True)
for i, row in df_time.iterrows():
    r = HR + 1 + i
    ws3.cell(row=r, column=1, value=row["label"])
    ws3.cell(row=r, column=2, value=row["inference_model"])
    for j, (_, col) in enumerate(time_cols, start=3):
        ws3.cell(row=r, column=j, value=float(row[col]))
    ws3.cell(row=r, column=3 + len(time_cols), value=float(row["mean_time"]))

end3 = HR + len(df_time)
style_body(ws3, HR + 1, end3, len(headers3))
for r in range(HR + 1, end3 + 1):
    for c in range(3, len(headers3) + 1):
        ws3.cell(row=r, column=c).number_format = "0.00"

for j in range(3, len(headers3) + 1):
    col_letter = get_column_letter(j)
    ws3.conditional_formatting.add(
        f"{col_letter}{HR+1}:{col_letter}{end3}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"))

auto_width(ws3, [32, 18] + [16] * len(time_cols) + [16])
ws3.freeze_panes = "C4"

# ──────────────────────────── Sheet 4: Reasoning vs Search ───────────────
ws4 = wb.create_sheet("Decomposição de Tempo")
ws4["A1"] = "Decomposição do Tempo Médio por Questão (s)"
ws4["A1"].font = Font(name=FONT, bold=True, size=16, color="1F4E78")
ws4.merge_cells("A1:F1")

headers4 = ["Experimento", "Inference", "Tempo Total", "Busca (RAG)", "Reasoning", "Outro"]
for i, h in enumerate(headers4, start=1):
    ws4.cell(row=HR, column=i, value=h)
style_header(ws4, HR, len(headers4))
ws4.row_dimensions[HR].height = 30

df_dec = df.sort_values("mean_time").reset_index(drop=True)
for i, row in df_dec.iterrows():
    r = HR + 1 + i
    total = float(row["mean_time"])
    search = float(row["mean_search_time"])
    reason = float(row["mean_reasoning_time"])
    other = max(total - search - reason, 0)
    ws4.cell(row=r, column=1, value=row["label"])
    ws4.cell(row=r, column=2, value=row["inference_model"])
    ws4.cell(row=r, column=3, value=total)
    ws4.cell(row=r, column=4, value=search)
    ws4.cell(row=r, column=5, value=reason)
    ws4.cell(row=r, column=6, value=other)

end4 = HR + len(df_dec)
style_body(ws4, HR + 1, end4, len(headers4))
for r in range(HR + 1, end4 + 1):
    for c in range(3, len(headers4) + 1):
        ws4.cell(row=r, column=c).number_format = "0.00"

auto_width(ws4, [32, 18, 14, 14, 14, 14])
ws4.freeze_panes = "C4"

wb.save(DST)
print(f"Saved: {DST}")
