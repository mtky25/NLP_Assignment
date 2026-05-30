"""Generate polished report v2 (English) with RAM/efficiency metrics, raw data, charts."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BubbleChart, Reference, Series, BarChart
from openpyxl.chart.label import DataLabelList

SRC = r"C:\Users\marce\OneDrive\Documentos\marcelo_final.xlsx"
DST = r"C:\Users\marce\OneDrive\Documentos\marcelo_final_report_v5.xlsx"

# ── Model sizes in MB (from `ollama list`), kept internally for math only ──
MODEL_MB = {
    "nomic-embed-text":      274,
    "nomic-embed-text:latest": 274,
    "qwen2.5:0.5b":          397,
    "qwen2-math:1.5b":       934,
    "llama3.2:1b":          1331,
    "qwen2.5:1.5b":          986,
    "llama3.2:3b":          2048,
    "qwen2.5:3b":           1946,
    "phi3.5:latest":        2253,
    "gemma3:4b":            3379,
    "gemma2:latest":        5530,
}
FIXED_MB = MODEL_MB["nomic-embed-text"] + MODEL_MB["qwen2.5:0.5b"] + MODEL_MB["qwen2-math:1.5b"]
FIXED_GB = FIXED_MB / 1024.0

W_ACC, W_TIME, W_RAM = 0.50, 0.25, 0.25

COLS = [
    "experiment_id", "username", "notes", "approach", "inference_model",
    "inference_model_size", "debug", "mode", "transcription_model", "is_rag",
    "embedding_model", "embedding_model_size",
    "mean_question_accuracy", "mean_time", "mean_search_time",
    "mean_reasoning_time", "mean_transcription_time",
    "ent_time", "ent_acc", "ent_search", "ent_reason", "ent_trans",
    "sci_time", "sci_acc", "sci_search", "sci_reason", "sci_trans",
    "anc_time", "anc_acc", "anc_search", "anc_reason", "anc_trans",
    "math_time", "math_acc", "math_search", "math_reason", "math_trans",
    "news_time", "news_acc", "news_search", "news_reason",
    "phil_time", "phil_acc", "phil_search", "phil_reason",
    "timestamp",
]

df = pd.read_excel(SRC, header=None, names=COLS)
df["label"] = df["notes"]
df["fallback_model"] = df["label"].str.split("/").str[1].str.strip()

def lookup_mb(name):
    if name in MODEL_MB: return MODEL_MB[name]
    base = name.split(":")[0]
    for k, v in MODEL_MB.items():
        if k.startswith(base): return v
    return 0

df["inf_gb"]   = df["inference_model"].apply(lookup_mb) / 1024.0
df["fb_gb"]    = df["fallback_model"].apply(lookup_mb)  / 1024.0
df["ram_gb"]   = df["inf_gb"] + df["fb_gb"] + FIXED_GB

df["acc_per_gb"]    = df["mean_question_accuracy"] / df["ram_gb"]
df["cost_gb_s"]     = df["mean_time"] * df["ram_gb"]
df["eff_composite"] = df["mean_question_accuracy"] / df["cost_gb_s"]

acc_min, acc_max = df["mean_question_accuracy"].min(), df["mean_question_accuracy"].max()
t_min,   t_max   = df["mean_time"].min(),               df["mean_time"].max()
r_min,   r_max   = df["ram_gb"].min(),                  df["ram_gb"].max()
df["acc_n"]  = (df["mean_question_accuracy"] - acc_min) / (acc_max - acc_min)
df["time_n"] = (df["mean_time"] - t_min)               / (t_max - t_min)
df["ram_n"]  = (df["ram_gb"] - r_min)                  / (r_max - r_min)
df["score"]  = W_ACC * df["acc_n"] - W_TIME * df["time_n"] - W_RAM * df["ram_n"]

def is_pareto(i):
    a, t, r = df.loc[i, "mean_question_accuracy"], df.loc[i, "mean_time"], df.loc[i, "ram_gb"]
    for j in df.index:
        if j == i: continue
        if (df.loc[j, "mean_question_accuracy"] >= a and
            df.loc[j, "mean_time"] <= t and df.loc[j, "ram_gb"] <= r and
            (df.loc[j, "mean_question_accuracy"] > a or
             df.loc[j, "mean_time"] < t or df.loc[j, "ram_gb"] < r)):
            return False
    return True
df["pareto"] = [is_pareto(i) for i in df.index]

# ── Styles ────────────────────────────────────────────────────────────────
FONT = "Arial"
HEADER_FILL  = PatternFill("solid", start_color="1F4E78")
BAND_FILL    = PatternFill("solid", start_color="F2F2F2")
PARETO_FILL  = PatternFill("solid", start_color="E2EFDA")
HEADER_FONT  = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name=FONT, size=10)
TITLE_FONT   = Font(name=FONT, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT= Font(name=FONT, italic=True, size=10, color="595959")
SECTION_FONT = Font(name=FONT, bold=True, size=13, color="1F4E78")
THIN  = Side(style="thin", color="BFBFBF")
BORDER= Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER

def style_body(ws, sr, er, ncols, band=True):
    for r in range(sr, er + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER
            if band and (r - sr) % 2 == 1:
                cell.fill = BAND_FILL

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def cs_higher_better(ws, rng):
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))

def cs_lower_better(ws, rng):
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B"))

wb = Workbook()

# ═════════════════════════ Sheet 1: Executive Summary ═════════════════════
ws = wb.active
ws.title = "Executive Summary"
ws["A1"] = "Final Report v2 — Model Comparison (Millionaire RAG)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:K1")
ws["A2"] = ("Embedding: nomic-embed-text  |  Translator: qwen2.5:0.5b  |  Math: qwen2-math:1.5b  |  "
            "10 games/theme  |  Score = 0.50·acc − 0.25·time − 0.25·RAM (normalised)")
ws["A2"].font = SUBTITLE_FONT
ws.merge_cells("A2:K2")
ws.row_dimensions[1].height = 24

summary = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
summary.insert(0, "rank", range(1, len(summary) + 1))

headers = ["#", "Experiment (inference / fallback)", "Inference Model",
           "Accuracy", "Time (s)", "RAM (GB)",
           "Acc / GB", "Time·GB", "Efficiency", "Score", "Pareto"]
HR = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HR, column=i, value=h)
style_header(ws, HR, len(headers))
ws.row_dimensions[HR].height = 30

sr = HR + 1
for i, row in summary.iterrows():
    r = sr + i
    ws.cell(row=r, column=1,  value=int(row["rank"]))
    ws.cell(row=r, column=2,  value=row["label"])
    ws.cell(row=r, column=3,  value=row["inference_model"])
    ws.cell(row=r, column=4,  value=float(row["mean_question_accuracy"]))
    ws.cell(row=r, column=5,  value=float(row["mean_time"]))
    ws.cell(row=r, column=6,  value=float(row["ram_gb"]))
    ws.cell(row=r, column=7,  value=float(row["acc_per_gb"]))
    ws.cell(row=r, column=8,  value=float(row["cost_gb_s"]))
    ws.cell(row=r, column=9,  value=float(row["eff_composite"]))
    ws.cell(row=r, column=10, value=float(row["score"]))
    ws.cell(row=r, column=11, value="✓" if row["pareto"] else "")

er = sr + len(summary) - 1
style_body(ws, sr, er, len(headers))

for r in range(sr, er + 1):
    if ws.cell(row=r, column=11).value == "✓":
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = PARETO_FILL

for r in range(sr, er + 1):
    ws.cell(row=r, column=4).number_format = "0.00%"
    ws.cell(row=r, column=5).number_format = "0.00"
    ws.cell(row=r, column=6).number_format = "0.00"
    ws.cell(row=r, column=7).number_format = "0.00%"
    ws.cell(row=r, column=8).number_format = "0.00"
    ws.cell(row=r, column=9).number_format = "0.0000"
    ws.cell(row=r, column=10).number_format = "0.000"

cs_higher_better(ws, f"D{sr}:D{er}")
cs_lower_better(ws,  f"E{sr}:E{er}")
cs_lower_better(ws,  f"F{sr}:F{er}")
cs_higher_better(ws, f"G{sr}:G{er}")
cs_lower_better(ws,  f"H{sr}:H{er}")
cs_higher_better(ws, f"I{sr}:I{er}")
cs_higher_better(ws, f"J{sr}:J{er}")

widths(ws, [5, 32, 16, 12, 11, 11, 11, 12, 13, 10, 9])

fr = er + 3
ws.cell(row=fr, column=1, value="Key Findings").font = SECTION_FONT
ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=11)

best_acc   = summary.iloc[0]
best_score = summary.sort_values("score", ascending=False).iloc[0]
best_eff   = summary.sort_values("eff_composite", ascending=False).iloc[0]
lightest   = summary.sort_values("ram_gb").iloc[0]
fastest    = summary.sort_values("mean_time").iloc[0]
pareto_lbls= ", ".join(summary[summary["pareto"]]["label"].tolist())

bullets = [
    f"• Highest accuracy: {best_acc['label']} ({best_acc['mean_question_accuracy']:.2%}) — uses {best_acc['ram_gb']:.2f} GB in {best_acc['mean_time']:.2f} s/question.",
    f"• Best composite score (50% acc / 25% time / 25% RAM): {best_score['label']} (score {best_score['score']:.3f}).",
    f"• Highest efficiency (acc / (time·GB)): {best_eff['label']} — {best_eff['eff_composite']:.4f}.",
    f"• Lightest in RAM: {lightest['label']} ({lightest['ram_gb']:.2f} GB).  Fastest: {fastest['label']} ({fastest['mean_time']:.2f} s).",
    f"• Pareto frontier (non-dominated on any axis): {pareto_lbls}.",
    f"• gemma2:latest as fallback costs +3.2 GB of RAM vs. phi3.5 and still loses 13.4 pp of accuracy — worst trade-off in the set.",
]
for i, b in enumerate(bullets):
    ws.cell(row=fr + 1 + i, column=1, value=b).font = BODY_FONT
    ws.merge_cells(start_row=fr + 1 + i, start_column=1, end_row=fr + 1 + i, end_column=11)

ws.freeze_panes = "C5"

# ═════════════════════════ Sheet 2: Models & RAM ══════════════════════════
ws2 = wb.create_sheet("Models & RAM")
ws2["A1"] = "RAM Composition by Experiment"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:H1")
ws2["A2"] = (f"Fixed models in every run: nomic-embed-text + qwen2.5:0.5b + qwen2-math:1.5b "
             f"= {FIXED_GB:.2f} GB")
ws2["A2"].font = SUBTITLE_FONT
ws2.merge_cells("A2:H2")

h2 = ["#", "Experiment", "Inference", "Inf. (GB)", "Fallback", "Fb. (GB)", "Fixed (GB)", "Total (GB)"]
HR = 4
for i, h in enumerate(h2, start=1):
    ws2.cell(row=HR, column=i, value=h)
style_header(ws2, HR, len(h2))
ws2.row_dimensions[HR].height = 30

df_ram = df.sort_values("ram_gb").reset_index(drop=True)
sr2 = HR + 1
for i, row in df_ram.iterrows():
    r = sr2 + i
    ws2.cell(row=r, column=1, value=i + 1)
    ws2.cell(row=r, column=2, value=row["label"])
    ws2.cell(row=r, column=3, value=row["inference_model"])
    ws2.cell(row=r, column=4, value=float(row["inf_gb"]))
    ws2.cell(row=r, column=5, value=row["fallback_model"])
    ws2.cell(row=r, column=6, value=float(row["fb_gb"]))
    ws2.cell(row=r, column=7, value=FIXED_GB)
    ws2.cell(row=r, column=8, value=float(row["ram_gb"]))

er2 = sr2 + len(df_ram) - 1
style_body(ws2, sr2, er2, len(h2))
for r in range(sr2, er2 + 1):
    for c in (4, 6, 7, 8):
        ws2.cell(row=r, column=c).number_format = "0.00"

cs_lower_better(ws2, f"H{sr2}:H{er2}")
widths(ws2, [5, 30, 16, 11, 16, 11, 11, 11])
ws2.freeze_panes = "C5"

# ═════════════════════════ Sheet 3: Trade-off ═════════════════════════════
ws3 = wb.create_sheet("Trade-off Acc x Time x RAM")
ws3["A1"] = "Trade-off Between Accuracy, Time and RAM"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:L1")
ws3["A2"] = "Green = best on each column. Light-green row = point on the Pareto frontier."
ws3["A2"].font = SUBTITLE_FONT
ws3.merge_cells("A2:L2")

h3 = ["Experiment", "Accuracy", "Time (s)", "RAM (GB)",
      "Acc (norm.)", "Time (norm.)", "RAM (norm.)",
      "Acc / GB", "Time·GB", "Efficiency", "Score", "Pareto"]
HR = 4
for i, h in enumerate(h3, start=1):
    ws3.cell(row=HR, column=i, value=h)
style_header(ws3, HR, len(h3))
ws3.row_dimensions[HR].height = 30

df_to = df.sort_values("score", ascending=False).reset_index(drop=True)
sr3 = HR + 1
for i, row in df_to.iterrows():
    r = sr3 + i
    ws3.cell(row=r, column=1,  value=row["label"])
    ws3.cell(row=r, column=2,  value=float(row["mean_question_accuracy"]))
    ws3.cell(row=r, column=3,  value=float(row["mean_time"]))
    ws3.cell(row=r, column=4,  value=float(row["ram_gb"]))
    ws3.cell(row=r, column=5,  value=float(row["acc_n"]))
    ws3.cell(row=r, column=6,  value=float(row["time_n"]))
    ws3.cell(row=r, column=7,  value=float(row["ram_n"]))
    ws3.cell(row=r, column=8,  value=float(row["acc_per_gb"]))
    ws3.cell(row=r, column=9,  value=float(row["cost_gb_s"]))
    ws3.cell(row=r, column=10, value=float(row["eff_composite"]))
    ws3.cell(row=r, column=11, value=float(row["score"]))
    ws3.cell(row=r, column=12, value="✓" if row["pareto"] else "")

er3 = sr3 + len(df_to) - 1
style_body(ws3, sr3, er3, len(h3))
for r in range(sr3, er3 + 1):
    if ws3.cell(row=r, column=12).value == "✓":
        for c in range(1, len(h3) + 1):
            ws3.cell(row=r, column=c).fill = PARETO_FILL

for r in range(sr3, er3 + 1):
    ws3.cell(row=r, column=2).number_format = "0.00%"
    ws3.cell(row=r, column=3).number_format = "0.00"
    ws3.cell(row=r, column=4).number_format = "0.00"
    for c in (5, 6, 7):
        ws3.cell(row=r, column=c).number_format = "0.000"
    ws3.cell(row=r, column=8).number_format  = "0.00%"
    ws3.cell(row=r, column=9).number_format  = "0.00"
    ws3.cell(row=r, column=10).number_format = "0.0000"
    ws3.cell(row=r, column=11).number_format = "0.000"

cs_higher_better(ws3, f"B{sr3}:B{er3}")
cs_lower_better(ws3,  f"C{sr3}:C{er3}")
cs_lower_better(ws3,  f"D{sr3}:D{er3}")
cs_higher_better(ws3, f"H{sr3}:H{er3}")
cs_lower_better(ws3,  f"I{sr3}:I{er3}")
cs_higher_better(ws3, f"J{sr3}:J{er3}")
cs_higher_better(ws3, f"K{sr3}:K{er3}")

widths(ws3, [30, 11, 11, 11, 12, 13, 12, 11, 11, 12, 10, 9])
ws3.freeze_panes = "B5"

# ═════════════════════════ Sheet 4: Accuracy by Theme ═════════════════════
ws4 = wb.create_sheet("Accuracy by Theme")
ws4["A1"] = "Accuracy by Theme"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:I1")

theme_cols = [("Entertainment","ent_acc"), ("Science & Nature","sci_acc"),
              ("Ancient History","anc_acc"), ("Maths","math_acc"),
              ("News","news_acc"), ("Philosophy & Psych.","phil_acc")]
h4 = ["Experiment", "Inference"] + [t[0] for t in theme_cols] + ["Overall"]
HR = 3
for i, h in enumerate(h4, start=1):
    ws4.cell(row=HR, column=i, value=h)
style_header(ws4, HR, len(h4))
ws4.row_dimensions[HR].height = 30

df_acc = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
for i, row in df_acc.iterrows():
    r = HR + 1 + i
    ws4.cell(row=r, column=1, value=row["label"])
    ws4.cell(row=r, column=2, value=row["inference_model"])
    for j, (_, col) in enumerate(theme_cols, start=3):
        ws4.cell(row=r, column=j, value=float(row[col]))
    ws4.cell(row=r, column=3 + len(theme_cols), value=float(row["mean_question_accuracy"]))

er4 = HR + len(df_acc)
style_body(ws4, HR + 1, er4, len(h4))
for r in range(HR + 1, er4 + 1):
    for c in range(3, len(h4) + 1):
        ws4.cell(row=r, column=c).number_format = "0.00%"
for j in range(3, len(h4) + 1):
    cs_higher_better(ws4, f"{get_column_letter(j)}{HR+1}:{get_column_letter(j)}{er4}")
widths(ws4, [30, 16] + [16] * len(theme_cols) + [12])
ws4.freeze_panes = "C4"

# ═════════════════════════ Sheet 5: Time by Theme ═════════════════════════
ws5 = wb.create_sheet("Time by Theme")
ws5["A1"] = "Mean Time (s) by Theme"
ws5["A1"].font = TITLE_FONT
ws5.merge_cells("A1:I1")

time_cols = [("Entertainment","ent_time"), ("Science & Nature","sci_time"),
             ("Ancient History","anc_time"), ("Maths","math_time"),
             ("News","news_time"), ("Philosophy & Psych.","phil_time")]
h5 = ["Experiment", "Inference"] + [t[0] for t in time_cols] + ["Total"]
HR = 3
for i, h in enumerate(h5, start=1):
    ws5.cell(row=HR, column=i, value=h)
style_header(ws5, HR, len(h5))
ws5.row_dimensions[HR].height = 30

df_t = df.sort_values("mean_time").reset_index(drop=True)
for i, row in df_t.iterrows():
    r = HR + 1 + i
    ws5.cell(row=r, column=1, value=row["label"])
    ws5.cell(row=r, column=2, value=row["inference_model"])
    for j, (_, col) in enumerate(time_cols, start=3):
        ws5.cell(row=r, column=j, value=float(row[col]))
    ws5.cell(row=r, column=3 + len(time_cols), value=float(row["mean_time"]))

er5 = HR + len(df_t)
style_body(ws5, HR + 1, er5, len(h5))
for r in range(HR + 1, er5 + 1):
    for c in range(3, len(h5) + 1):
        ws5.cell(row=r, column=c).number_format = "0.00"
for j in range(3, len(h5) + 1):
    cs_lower_better(ws5, f"{get_column_letter(j)}{HR+1}:{get_column_letter(j)}{er5}")
widths(ws5, [30, 16] + [16] * len(time_cols) + [12])
ws5.freeze_panes = "C4"

# ═════════════════════════ Sheet 6: Raw Data ══════════════════════════════
ws6 = wb.create_sheet("Raw Data")
ws6["A1"] = "Raw Data (formatted)"
ws6["A1"].font = TITLE_FONT
ws6.merge_cells("A1:AT1")

raw_headers = [
    "Experiment ID", "Username", "Notes (label)", "Approach", "Inference Model",
    "Inference Size", "Debug", "Mode", "Transcription Model", "Is RAG",
    "Embedding Model", "Embedding Size",
    "Mean Accuracy", "Mean Time (s)", "Mean Search (s)", "Mean Reasoning (s)", "Mean Transcription (s)",
    "Ent Time (s)", "Ent Acc.", "Ent Search (s)", "Ent Reason (s)", "Ent Trans (s)",
    "Sci Time (s)", "Sci Acc.", "Sci Search (s)", "Sci Reason (s)", "Sci Trans (s)",
    "Anc Time (s)", "Anc Acc.", "Anc Search (s)", "Anc Reason (s)", "Anc Trans (s)",
    "Math Time (s)", "Math Acc.", "Math Search (s)", "Math Reason (s)", "Math Trans (s)",
    "News Time (s)", "News Acc.", "News Search (s)", "News Reason (s)",
    "Phil Time (s)", "Phil Acc.", "Phil Search (s)", "Phil Reason (s)",
    "Timestamp",
]
HR = 3
for i, h in enumerate(raw_headers, start=1):
    ws6.cell(row=HR, column=i, value=h)
style_header(ws6, HR, len(raw_headers))
ws6.row_dimensions[HR].height = 32

ACC_COLS  = [13, 19, 24, 29, 34, 39, 43]
TIME_COLS = [14, 15, 16, 17, 18, 20, 21, 22, 23, 25, 26, 27,
             28, 30, 31, 32, 33, 35, 36, 37, 38, 40, 41, 42, 44, 45]

sr6 = HR + 1
for i, row in df.iterrows():
    r = sr6 + i
    vals = [row[c] for c in COLS]
    for j, v in enumerate(vals, start=1):
        ws6.cell(row=r, column=j, value=v)

er6 = sr6 + len(df) - 1
style_body(ws6, sr6, er6, len(raw_headers))

for r in range(sr6, er6 + 1):
    for c in ACC_COLS:
        ws6.cell(row=r, column=c).number_format = "0.00%"
    for c in TIME_COLS:
        cell = ws6.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"

raw_widths = [22, 10, 28, 10, 16, 10, 8, 8, 16, 8, 22, 12] + [14] * (len(raw_headers) - 13) + [20]
widths(ws6, raw_widths)
ws6.freeze_panes = "D4"

# ═════════════════════════ Sheet 7: All Metrics ══════════════════════════
ws_all = wb.create_sheet("All Metrics")
ws_all["A1"] = "All Metrics (raw + derived)"
ws_all["A1"].font = TITLE_FONT
ws_all.merge_cells("A1:AZ1")
ws_all["A2"] = ("Every metric for every experiment: raw fields from the benchmark plus computed metrics "
                "(RAM, normalised values, efficiency, composite score, Pareto flag).")
ws_all["A2"].font = SUBTITLE_FONT
ws_all.merge_cells("A2:AZ2")

all_headers = [
    # Identity
    "Experiment", "Inference Model", "Fallback Model", "Embedding Model",
    # Overall raw
    "Accuracy", "Mean Time (s)", "Mean Search (s)", "Mean Reasoning (s)", "Mean Transcription (s)",
    # Per-theme accuracy
    "Ent Acc.", "Sci Acc.", "Anc Acc.", "Math Acc.", "News Acc.", "Phil Acc.",
    # Per-theme time
    "Ent Time (s)", "Sci Time (s)", "Anc Time (s)", "Math Time (s)", "News Time (s)", "Phil Time (s)",
    # Per-theme search
    "Ent Search (s)", "Sci Search (s)", "Anc Search (s)", "Math Search (s)", "News Search (s)", "Phil Search (s)",
    # Per-theme reasoning
    "Ent Reason (s)", "Sci Reason (s)", "Anc Reason (s)", "Math Reason (s)", "News Reason (s)", "Phil Reason (s)",
    # RAM
    "Inference RAM (GB)", "Fallback RAM (GB)", "Fixed RAM (GB)", "Total RAM (GB)",
    # Derived
    "Acc / GB", "Time·GB (cost)", "Efficiency (acc/(t·GB))",
    "Acc (norm.)", "Time (norm.)", "RAM (norm.)", "Composite Score", "Pareto",
]
HR = 4
for i, h in enumerate(all_headers, start=1):
    ws_all.cell(row=HR, column=i, value=h)
style_header(ws_all, HR, len(all_headers))
ws_all.row_dimensions[HR].height = 36

df_all = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
sra = HR + 1
for i, row in df_all.iterrows():
    r = sra + i
    vals = [
        row["label"], row["inference_model"], row["fallback_model"], row["embedding_model"],
        float(row["mean_question_accuracy"]), float(row["mean_time"]),
        float(row["mean_search_time"]), float(row["mean_reasoning_time"]), float(row["mean_transcription_time"]),
        float(row["ent_acc"]), float(row["sci_acc"]), float(row["anc_acc"]),
        float(row["math_acc"]), float(row["news_acc"]), float(row["phil_acc"]),
        float(row["ent_time"]), float(row["sci_time"]), float(row["anc_time"]),
        float(row["math_time"]), float(row["news_time"]), float(row["phil_time"]),
        float(row["ent_search"]), float(row["sci_search"]), float(row["anc_search"]),
        float(row["math_search"]), float(row["news_search"]), float(row["phil_search"]),
        float(row["ent_reason"]), float(row["sci_reason"]), float(row["anc_reason"]),
        float(row["math_reason"]), float(row["news_reason"]), float(row["phil_reason"]),
        float(row["inf_gb"]), float(row["fb_gb"]), float(FIXED_GB), float(row["ram_gb"]),
        float(row["acc_per_gb"]), float(row["cost_gb_s"]), float(row["eff_composite"]),
        float(row["acc_n"]), float(row["time_n"]), float(row["ram_n"]),
        float(row["score"]), "✓" if row["pareto"] else "",
    ]
    for j, v in enumerate(vals, start=1):
        ws_all.cell(row=r, column=j, value=v)

era = sra + len(df_all) - 1
style_body(ws_all, sra, era, len(all_headers))

# Column index map (1-based) for formatting
PCT_COLS_ALL  = [5, 10, 11, 12, 13, 14, 15, 38]                    # accuracies + Acc/GB
NUM_COLS_ALL  = list(range(6, 10)) + list(range(16, 34)) + [34, 35, 36, 37, 39]  # times + RAM + cost
SMALL_COLS    = [40]                                                # efficiency (4 decimals)
NORM_COLS     = [41, 42, 43, 44]                                    # normalised + score

for r in range(sra, era + 1):
    for c in PCT_COLS_ALL:
        ws_all.cell(row=r, column=c).number_format = "0.00%"
    for c in NUM_COLS_ALL:
        ws_all.cell(row=r, column=c).number_format = "0.00"
    for c in SMALL_COLS:
        ws_all.cell(row=r, column=c).number_format = "0.0000"
    for c in NORM_COLS:
        ws_all.cell(row=r, column=c).number_format = "0.000"

# Pareto row highlight
for r in range(sra, era + 1):
    if ws_all.cell(row=r, column=len(all_headers)).value == "✓":
        for c in range(1, len(all_headers) + 1):
            ws_all.cell(row=r, column=c).fill = PARETO_FILL

# Heat-maps on key columns
cs_higher_better(ws_all, f"E{sra}:E{era}")   # Accuracy
cs_lower_better(ws_all,  f"F{sra}:F{era}")   # Mean Time
cs_lower_better(ws_all,  f"{get_column_letter(37)}{sra}:{get_column_letter(37)}{era}")  # Total RAM
cs_higher_better(ws_all, f"{get_column_letter(38)}{sra}:{get_column_letter(38)}{era}")  # Acc/GB
cs_lower_better(ws_all,  f"{get_column_letter(39)}{sra}:{get_column_letter(39)}{era}")  # Time·GB
cs_higher_better(ws_all, f"{get_column_letter(40)}{sra}:{get_column_letter(40)}{era}")  # Efficiency
cs_higher_better(ws_all, f"{get_column_letter(44)}{sra}:{get_column_letter(44)}{era}")  # Score

# Widths
all_widths = [30, 16, 16, 22] + [12] * (len(all_headers) - 4)
widths(ws_all, all_widths)
ws_all.freeze_panes = "E5"

# ═════════════════════════ Sheet 8: Charts ════════════════════════════════
ws7 = wb.create_sheet("Charts")
ws7["A1"] = "Visualisations"
ws7["A1"].font = TITLE_FONT
ws7.merge_cells("A1:N1")
ws7["A2"] = "Bubble chart — X = mean time (s), Y = accuracy, bubble size = RAM (GB)."
ws7["A2"].font = SUBTITLE_FONT
ws7.merge_cells("A2:N2")

# Hidden data block for bubble chart
DATA_START = 200
ws7.cell(row=DATA_START, column=1, value="label")
ws7.cell(row=DATA_START, column=2, value="time")
ws7.cell(row=DATA_START, column=3, value="accuracy")
ws7.cell(row=DATA_START, column=4, value="ram_gb")
df_chart = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
for i, row in df_chart.iterrows():
    r = DATA_START + 1 + i
    ws7.cell(row=r, column=1, value=row["label"])
    ws7.cell(row=r, column=2, value=float(row["mean_time"]))
    ws7.cell(row=r, column=3, value=float(row["mean_question_accuracy"]))
    ws7.cell(row=r, column=4, value=float(row["ram_gb"]))

from openpyxl.chart import ScatterChart
# Use ScatterChart (much more stable in openpyxl/Excel) — RAM shown in legend table
scatter = ScatterChart()
scatter.title  = "Trade-off: Accuracy vs. Time (RAM in legend table)"
scatter.x_axis.title = "Mean time per question (s)"
scatter.y_axis.title = "Overall accuracy"
scatter.height = 14
scatter.width  = 24
scatter.legend = None

n = len(df_chart)
xref = Reference(ws7, min_col=2, min_row=DATA_START + 1, max_row=DATA_START + n)
yref = Reference(ws7, min_col=3, min_row=DATA_START + 1, max_row=DATA_START + n)
series = Series(values=yref, xvalues=xref, title="Experiments")
scatter.series.append(series)

ws7.add_chart(scatter, "A4")

# Reference legend next to the chart — maps each experiment to its (x, y, size)
LEG_START = 4
ws7.cell(row=LEG_START, column=16, value="Experiment").font = HEADER_FONT
ws7.cell(row=LEG_START, column=17, value="Time (s)").font  = HEADER_FONT
ws7.cell(row=LEG_START, column=18, value="Accuracy").font  = HEADER_FONT
ws7.cell(row=LEG_START, column=19, value="RAM (GB)").font  = HEADER_FONT
for c in range(16, 20):
    cell = ws7.cell(row=LEG_START, column=c)
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
for i, row in df_chart.iterrows():
    rr = LEG_START + 1 + i
    ws7.cell(row=rr, column=16, value=row["label"])
    ws7.cell(row=rr, column=17, value=float(row["mean_time"])).number_format = "0.00"
    ws7.cell(row=rr, column=18, value=float(row["mean_question_accuracy"])).number_format = "0.00%"
    ws7.cell(row=rr, column=19, value=float(row["ram_gb"])).number_format = "0.00"
    for c in range(16, 20):
        cell = ws7.cell(row=rr, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = LEFT if c == 16 else CENTER
        if i % 2 == 1:
            cell.fill = BAND_FILL
for col, w in zip([16, 17, 18, 19], [30, 11, 11, 11]):
    ws7.column_dimensions[get_column_letter(col)].width = w

# ── Bar chart: Composite Score ───────────────────────────────────────────
BAR_START = 230
ws7.cell(row=BAR_START, column=1, value="Experiment")
ws7.cell(row=BAR_START, column=2, value="Score")
df_bar = df.sort_values("score", ascending=False).reset_index(drop=True)
for i, row in df_bar.iterrows():
    ws7.cell(row=BAR_START + 1 + i, column=1, value=row["label"])
    ws7.cell(row=BAR_START + 1 + i, column=2, value=float(row["score"]))

bar = BarChart()
bar.type = "bar"
bar.title = "Composite Score (0.50·acc − 0.25·time − 0.25·RAM, normalised)"
bar.y_axis.title = "Experiment"
bar.x_axis.title = "Score"
bar.height = 14
bar.width  = 24
bar.legend = None

data_ref = Reference(ws7, min_col=2, min_row=BAR_START, max_row=BAR_START + len(df_bar))
cats_ref = Reference(ws7, min_col=1, min_row=BAR_START + 1, max_row=BAR_START + len(df_bar))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
ws7.add_chart(bar, "A42")

# ── Bar chart: Overall Accuracy ──────────────────────────────────────────
BAR2_START = 260
ws7.cell(row=BAR2_START, column=1, value="Experiment")
ws7.cell(row=BAR2_START, column=2, value="Accuracy")
df_acc2 = df.sort_values("mean_question_accuracy", ascending=False).reset_index(drop=True)
for i, row in df_acc2.iterrows():
    ws7.cell(row=BAR2_START + 1 + i, column=1, value=row["label"])
    ws7.cell(row=BAR2_START + 1 + i, column=2, value=float(row["mean_question_accuracy"]))

bar2 = BarChart()
bar2.type = "bar"
bar2.title = "Overall Accuracy by Experiment"
bar2.y_axis.title = "Experiment"
bar2.x_axis.title = "Accuracy"
bar2.height = 14
bar2.width  = 24
bar2.legend = None

data_ref2 = Reference(ws7, min_col=2, min_row=BAR2_START, max_row=BAR2_START + len(df_acc2))
cats_ref2 = Reference(ws7, min_col=1, min_row=BAR2_START + 1, max_row=BAR2_START + len(df_acc2))
bar2.add_data(data_ref2, titles_from_data=True)
bar2.set_categories(cats_ref2)
ws7.add_chart(bar2, "A76")

wb.save(DST)
print(f"Saved: {DST}")
